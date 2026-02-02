from openpyxl import load_workbook
from datetime import date
from pathlib import Path

# -------------------------------------------------
# CONFIG
# -------------------------------------------------
EXCEL_PATH = Path("data/Job Application Tracker.csv")
SHEET_NAME = "Master"
TABLE_NAME = "ApplicationTracker"

WRITABLE_FIELDS = {
    "Date",
    "Status",
    "Job Title",
    "Company",
    "Office",
    "Location",
    "Salaried",
    "Low",
    "High",
    "Likely",
    "Want",
    "Updated",
    "Source",
    "Link",
}

REQUIRED_FIELDS = {"Job Title", "Company"}

# -------------------------------------------------
# INPUT
# -------------------------------------------------
entry = {
    "Job Title": "Test Apprentice",
    "Company": "Fake",
    "Office": "Hybrid",
    "Location": "Lynchburg, VA",
    "Salaried": "Yes",
    "Low": 39213,
    "High": 49232,
    "Likely": 0,
    "Want": 9,
    "Source": "Company Site",
    "Link": "www.linkedin.com",
}

# -------------------------------------------------
# AUTO-FILL
# -------------------------------------------------
today = date.today()
entry.setdefault("Date", today)
entry.setdefault("Updated", today)
entry.setdefault("Status", "Applied")

# -------------------------------------------------
# VALIDATION
# -------------------------------------------------
if not EXCEL_PATH.exists():
    raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

missing = REQUIRED_FIELDS - entry.keys()
if missing:
    raise ValueError(f"Missing required fields: {missing}")

illegal_fields = set(entry.keys()) - WRITABLE_FIELDS
if illegal_fields:
    raise ValueError(f"Attempt to write protected columns: {illegal_fields}")

# -------------------------------------------------
# LOAD
# -------------------------------------------------
wb = load_workbook(EXCEL_PATH)
if SHEET_NAME not in wb.sheetnames:
    raise ValueError(f"Sheet not found: {SHEET_NAME}. Available: {wb.sheetnames}")

ws = wb[SHEET_NAME]

# -------------------------------------------------
# HELPERS
# -------------------------------------------------
def find_header_row(worksheet, must_have=("Job Title", "Company"), max_scan=50):
    for r in range(1, max_scan + 1):
        values = [c.value for c in worksheet[r]]
        if all(h in values for h in must_have):
            return r
    raise ValueError("Could not find the header row (missing expected headers).")

def extend_table_to_row(worksheet, table_name: str, new_last_row: int):
    if table_name not in worksheet.tables:
        raise ValueError(
            f"Table '{table_name}' not found. "
            f"Available tables: {list(worksheet.tables.keys())}"
        )

    table = worksheet.tables[table_name]
    start_cell, end_cell = table.ref.split(":")
    start_col = "".join(c for c in start_cell if c.isalpha())
    start_row = int("".join(c for c in start_cell if c.isdigit()))
    end_col = "".join(c for c in end_cell if c.isalpha())

    table.ref = f"{start_col}{start_row}:{end_col}{new_last_row}"

# -------------------------------------------------
# MAP HEADERS
# -------------------------------------------------
header_row = find_header_row(ws)

headers = {
    cell.value: cell.column
    for cell in ws[header_row]
    if cell.value is not None
}

unknown_fields = set(entry.keys()) - set(headers.keys())
if unknown_fields:
    raise ValueError(f"Unknown columns in entry: {unknown_fields}")

# -------------------------------------------------
# FIND NEXT EMPTY ROW
# -------------------------------------------------
company_col = headers["Company"]
row = header_row + 1
while ws.cell(row=row, column=company_col).value not in (None, ""):
    row += 1

# -------------------------------------------------
# WRITE
# -------------------------------------------------
for field, value in entry.items():
    ws.cell(row=row, column=headers[field], value=value)

# -------------------------------------------------
# EXTEND APPLICATIONTRACKER TABLE
# -------------------------------------------------
extend_table_to_row(ws, TABLE_NAME, row)

# -------------------------------------------------
# SAVE
# -------------------------------------------------
wb.save(EXCEL_PATH)

print(f"✅ Application appended to row {row} and table '{TABLE_NAME}' extended.")
